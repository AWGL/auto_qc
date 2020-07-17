import openpyxl
from ..models import RunAnalysis

def make_kpi_excel(runs):
    # empty list to save required run info
    all_runs_list = []

    # loop through runs, save required data to list
    for run in runs:

        # query all analyses linked to run
        run_id = run.run_id
        run_analyses = RunAnalysis.objects.filter(run=run_id)
        

        # pull out the worksheets and panels for each run
        panel = []
        worksheets = []
        for analysis in run_analyses:
            panel.append(analysis.analysis_type.analysis_type_id)
            worksheets.append(analysis.get_worksheets())

        # pull out remaining data from run
        all_runs_list.append([
            '|'.join(set(panel)),
            run_id,
            '|'.join(set(worksheets)),
            '',     # placeholder for empty field in excel sheet
            run.setup_date,
            run.instrument_date
        ])

    # ------------------------------------------------------------------------
    # make excel workbook
    wb = openpyxl.Workbook()

    # make excel sheet for all runs
    ws = wb.active
    ws.title = 'all'

    # make list of headers for each tab
    headers = ['Panel', 'Run ID', 'Worksheet', 'Worksheet Date', 'Setup Date', 'Run Date']
    row_num = 1

    # write headers
    for col, header in enumerate(headers):
        c = ws.cell(row=row_num, column=col+1)
        c.value = header
    row_num += 1

    # write runs
    for run in all_runs_list:
        for col, r in enumerate(run):
            c = ws.cell(row=row_num, column=col+1)
            c.value = r
        row_num += 1

    # ------------------------------------------------------------------------
    # make excel sheet for each panel in panels list

    # add extra headers
    headers = headers + ['TAT1', 'TAT2', 'Total TAT']

    # dictionary of panels, key is panel ID in database, value is title for the panel's tab
    panels = {
        'NGHS-101X': 'CRM',
        'NGHS-102X': 'BRCA', 
        'SMP2v3': 'CRUK', 
        'NIPT': 'NIPT', 
        'TruSightMyeloid': 'Myeloid', 
        'RochePanCancer': 'PanCan', 
        'IlluminaTruSightOne': 'TSO', 
        'IlluminaTruSightCancer': 'TSC', 
        'NexteraDNAFlex': 'WGS', 
        'AgilentOGTFH': 'FH'
    }
    
    # create sheet for each panel in panel dictionary
    for panel in panels:
        ws = wb.create_sheet(title=panels[panel])
        row_num = 1

        # write headers
        for col, header in enumerate(headers):
            c = ws.cell(row=row_num, column=col+1)
            c.value = header
        row_num += 1

        # check if run is in the panel, write to tab if it is
        for run in all_runs_list:
            if panel in run[0]:
                for col, r in enumerate(run):
                    c = ws.cell(row=row_num, column=col+1)
                    c.value = r

                # add excel formulae for calculating TATs
                # TAT1 - worksheet date -> setup date
                c = ws.cell(row=row_num, column=col+2)
                c.value = f'=NETWORKDAYS(D{row_num},E{row_num},1)'

                # TAT2 - setp date -> run date
                c = ws.cell(row=row_num, column=col+3)
                c.value = f'=NETWORKDAYS(E{row_num},F{row_num},1)'

                # Total TAT - worksheet date -> run date
                c = ws.cell(row=row_num, column=col+4)
                c.value = f'=NETWORKDAYS(D{row_num},F{row_num},1)'

                row_num += 1


    # ------------------------------------------------------------------------
    # other - tab for any run that doesnt fall within a panel
    ws = wb.create_sheet(title='other')
    row_num = 1

    # write headers
    for col, header in enumerate(headers):
        c = ws.cell(row=row_num, column=col+1)
        c.value = header
    row_num += 1

    # write runs
    for run in all_runs_list:
        # check if run is in one of the panels
        in_panel = False
        for panel in panels:
            if panel in run[0]:
                in_panel = True

        # if it isn't, add it to the 'other' tab
        if not in_panel:
            for col, r in enumerate(run):
                c = ws.cell(row=row_num, column=col+1)
                c.value = r

            # add excel formulae for calculating TATs
            # TAT1 - worksheet date -> setup date
            c = ws.cell(row=row_num, column=col+2)
            c.value = f'=NETWORKDAYS(D{row_num},E{row_num},1)'

            # TAT2 - setp date -> run date
            c = ws.cell(row=row_num, column=col+3)
            c.value = f'=NETWORKDAYS(E{row_num},F{row_num},1)'

            # Total TAT - worksheet date -> run date
            c = ws.cell(row=row_num, column=col+4)
            c.value = f'=NETWORKDAYS(D{row_num},F{row_num},1)'
            row_num += 1


    return wb